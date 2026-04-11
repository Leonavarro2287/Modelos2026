# modelos_decision/analizador.py
from Modelos2026.normalizador import ModelosNormalizacion
import pandas as pd
import numpy as np
from google.colab import files
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

def analizar():
    print("\n" + "="*80)
    print("ANÁLISIS DE NORMALIZACIÓN Y PONDERACIÓN")
    print("="*80)

    print("\n📂 Por favor, suba su archivo Excel:")
    uploaded = files.upload()
    filename = list(uploaded.keys())[0]
    df_original = pd.read_excel(uploaded[filename])

    print(f"\n✅ Archivo cargado: {filename}")
    print(f"📊 Dimensiones: {df_original.shape[0]} filas, {df_original.shape[1]} columnas")

    print("\n📋 Primeras 5 filas:")
    print(df_original.head())

    print("\n" + "="*80)
    print("🔍 IDENTIFICACIÓN DE ALTERNATIVAS")
    print("="*80)
    columna_id = input("Nombre de la columna identificadora: ").strip()

    if columna_id == "":
        columna_id = "Alternativa"
        df_original[columna_id] = [f"Alt_{i+1}" for i in range(len(df_original))]

    if columna_id not in df_original.columns:
        columna_id = df_original.columns[0]

    print(f"✅ Columna identificadora: '{columna_id}'")

    print("\n📋 COLUMNAS DISPONIBLES:")
    for i, col in enumerate(df_original.columns):
        print(f"   {i}: '{col}'")

    print("\n" + "="*80)
    print("🔢 CRITERIOS")
    print("="*80)
    criterios_input = input("Columnas criterio (separadas por coma): ").split(',')
    criterios = [c.strip() for c in criterios_input]

    criterios_validos = []
    for col in criterios:
        if col == columna_id:
            continue
        if col in df_original.columns:
            criterios_validos.append(col)
    criterios = criterios_validos

    if len(criterios) == 0:
        criterios = [col for col in df_original.select_dtypes(include=['number']).columns if col != columna_id]

    print(f"\n✅ Criterios: {criterios}")

    print("\n" + "="*80)
    print("📈 SENTIDO (MAX o MIN)")
    print("="*80)
    minimos = []
    for col in criterios:
        while True:
            sentido = input(f"   {col} (MAX/MIN): ").strip().upper()
            if sentido in ['MAX', 'MIN']:
                if sentido == 'MIN':
                    minimos.append(col)
                break

    print(f"\n✅ Criterios MIN: {minimos}")

    print("\n" + "="*80)
    print("TRANSFORMACIONES")
    print("="*80)

    df_trabajo = df_original[[columna_id] + criterios].copy()
    df_transformado = df_trabajo.copy()

    for col in minimos:
        if (df_transformado[col] == 0).any():
            df_transformado[col] = -df_transformado[col]
        else:
            df_transformado[col] = 1 / df_transformado[col]

    for col in criterios:
        if (df_transformado[col] < 0).any():
            min_val = df_transformado[col].min()
            constante = abs(min_val) + 1
            df_transformado[col] = df_transformado[col] + constante

    print("\n📊 Datos transformados:")
    print(df_transformado.head())

    print("\n" + "="*80)
    print("NORMALIZACIÓN")
    print("="*80)

    motor = ModelosNormalizacion()
    df_criterios = df_transformado[criterios]

    Referencias = {}
    for col in criterios:
        Referencias[col] = [
            df_criterios[col].min(),
            df_criterios[col].max(),
            df_criterios[col].quantile(0.75),
            df_criterios[col].max()
        ]

    resultados = motor.ejecutar_todo(
        df=df_criterios,
        minimo=[],
        metas_rim=Referencias,
        n_intervalos_oecd=4
    )

    print(f"Métodos disponibles: {list(resultados.keys())}")
    matrices_norm = list(resultados.keys())

    print("\n" + "="*80)
    print("CONFIGURACIÓN DE PONDERACIONES")
    print("="*80)

    metodos_ponderacion = ['Uniforme', 'Desvio_Estandar', 'Entropia', 'CRITIC']
    norm_elegida = {}

    for metodo in metodos_ponderacion:
        print(f"\n🔹 {metodo}")
        for i, m in enumerate(matrices_norm):
            print(f"   {i+1}. {m}")
        while True:
            try:
                opcion = int(input("   Número: ")) - 1
                if 0 <= opcion < len(matrices_norm):
                    norm_elegida[metodo] = matrices_norm[opcion]
                    break
            except ValueError:
                print("   Ingrese un número.")

    print("\n" + "="*80)
    print("CALCULANDO PESOS")
    print("="*80)

    alternativas = df_transformado[columna_id].tolist()
    resultados_ponderacion = {}

    # Uniforme
    matriz_uni = resultados[norm_elegida['Uniforme']][criterios].copy()
    n = len(criterios)
    w_uniforme = np.array([1/n] * n)
    puntaje_uni = matriz_uni.dot(w_uniforme).values
    ranking_uni = pd.Series(puntaje_uni).rank(ascending=False).astype(int)
    resultados_ponderacion['Uniforme'] = {
        'pesos': w_uniforme, 'puntajes': puntaje_uni, 'ranking': ranking_uni,
        'normalizacion': norm_elegida['Uniforme']
    }

    # Desvío Estándar
    matriz_ds = resultados[norm_elegida['Desvio_Estandar']][criterios].copy()
    sigma = matriz_ds.std()
    w_ds = sigma / sigma.sum()
    puntaje_ds = matriz_ds.dot(w_ds).values
    ranking_ds = pd.Series(puntaje_ds).rank(ascending=False).astype(int)
    resultados_ponderacion['Desvio_Estandar'] = {
        'pesos': w_ds.values, 'puntajes': puntaje_ds, 'ranking': ranking_ds,
        'normalizacion': norm_elegida['Desvio_Estandar']
    }

    # Entropía
    matriz_ent = resultados[norm_elegida['Entropia']][criterios].copy()
    m = len(matriz_ent)
    k = 1 / np.log(m)
    entropias = {}
    for col in criterios:
        p = matriz_ent[col] * np.log(matriz_ent[col] + 1e-12)
        e_j = -k * p.sum()
        entropias[col] = 1 - e_j
    suma_div = sum(entropias.values())
    w_ent = np.array([entropias[col] / suma_div for col in criterios])
    puntaje_ent = matriz_ent.dot(w_ent).values
    ranking_ent = pd.Series(puntaje_ent).rank(ascending=False).astype(int)
    resultados_ponderacion['Entropia'] = {
        'pesos': w_ent, 'puntajes': puntaje_ent, 'ranking': ranking_ent,
        'normalizacion': norm_elegida['Entropia']
    }

    # CRITIC
    matriz_crit = resultados[norm_elegida['CRITIC']][criterios].copy()
    corr = matriz_crit.corr()
    sigma_crit = matriz_crit.std()
    suma_conflicto = {}
    for col in criterios:
        suma_conflicto[col] = (1 - corr[col]).sum()
    C_j = {col: sigma_crit[col] * suma_conflicto[col] for col in criterios}
    suma_C = sum(C_j.values())
    w_crit = np.array([C_j[col] / suma_C for col in criterios])
    puntaje_crit = matriz_crit.dot(w_crit).values
    ranking_crit = pd.Series(puntaje_crit).rank(ascending=False).astype(int)
    resultados_ponderacion['CRITIC'] = {
        'pesos': w_crit, 'puntajes': puntaje_crit, 'ranking': ranking_crit,
        'normalizacion': norm_elegida['CRITIC']
    }

    # Estadísticas
    estadisticas = []
    for col in criterios:
        datos = df_transformado[col]
        stats = {
            'Criterio': col,
            'Media': round(datos.mean(), 4),
            'Mediana': round(datos.median(), 4),
            'Mínimo': round(datos.min(), 4),
            'Máximo': round(datos.max(), 4),
            'Rango': round(datos.max() - datos.min(), 4),
            'Desvío Estándar': round(datos.std(), 4),
            'Q1 (25%)': round(datos.quantile(0.25), 4),
            'Q3 (75%)': round(datos.quantile(0.75), 4)
        }
        estadisticas.append(stats)
    df_estadisticas = pd.DataFrame(estadisticas)

    # Rankings
    df_rankings = pd.DataFrame()
    df_rankings[columna_id] = alternativas
    for metodo, datos in resultados_ponderacion.items():
        df_rankings[f'Ranking_{metodo}'] = datos['ranking']

    # Normalizaciones juntas
    normalizaciones_lista = []
    for nombre, matriz in resultados.items():
        matriz_con_id = matriz.copy()
        matriz_con_id.insert(0, columna_id, alternativas)
        for col in matriz_con_id.select_dtypes(include=['float64', 'float32']).columns:
            matriz_con_id[col] = matriz_con_id[col].round(4)
        df_titulo = pd.DataFrame([[f"NORMALIZACIÓN: {nombre}"]], columns=[columna_id])
        normalizaciones_lista.append(df_titulo)
        normalizaciones_lista.append(matriz_con_id)
        normalizaciones_lista.append(pd.DataFrame([[]]))
    df_normalizaciones = pd.concat(normalizaciones_lista, ignore_index=True)

    # Ponderaciones juntas
    ponderaciones_lista = []
    for metodo, datos in resultados_ponderacion.items():
        df_titulo = pd.DataFrame([[f"MÉTODO DE PONDERACIÓN: {metodo}"]], columns=['A'])
        ponderaciones_lista.append(df_titulo)
        ponderaciones_lista.append(pd.DataFrame([[]]))
        df_sub_pesos = pd.DataFrame([["PESOS DE CRITERIOS"]], columns=['A'])
        ponderaciones_lista.append(df_sub_pesos)
        df_pesos = pd.DataFrame({'Criterio': criterios, 'Peso': np.round(datos['pesos'], 4)})
        ponderaciones_lista.append(df_pesos)
        ponderaciones_lista.append(pd.DataFrame([[]]))
        df_sub_puntajes = pd.DataFrame([["PUNTAJES Y RANKINGS"]], columns=['A'])
        ponderaciones_lista.append(df_sub_puntajes)
        df_puntajes = pd.DataFrame({
            columna_id: alternativas,
            'Puntaje': np.round(datos['puntajes'], 4),
            'Ranking': datos['ranking']
        })
        ponderaciones_lista.append(df_puntajes)
        ponderaciones_lista.append(pd.DataFrame([[]]))
        df_nota = pd.DataFrame([[f"Normalización usada: {datos['normalizacion']}"]], columns=['A'])
        ponderaciones_lista.append(df_nota)
        ponderaciones_lista.append(pd.DataFrame([[]]))
        ponderaciones_lista.append(pd.DataFrame([[]]))
    df_ponderaciones = pd.concat(ponderaciones_lista, ignore_index=True)

    # Exportar
    output_excel = "Informe_Modelos_de_Decision_2026.xlsx"

    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        df_original.to_excel(writer, sheet_name='1_Datos_Originales', index=False)
        df_transformado.to_excel(writer, sheet_name='2_Datos_Transformados', index=False)
        df_estadisticas.to_excel(writer, sheet_name='3_Estadisticas', index=False)
        df_normalizaciones.to_excel(writer, sheet_name='4_Normalizaciones', index=False, header=False)
        df_ponderaciones.to_excel(writer, sheet_name='5_Ponderaciones', index=False, header=False)
        df_rankings.to_excel(writer, sheet_name='6_Rankings_Comparativos', index=False)

    from openpyxl import load_workbook
    wb = load_workbook(output_excel)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = '0.0000'
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = min(cell_length, 50)
                except:
                    pass
            adjusted_width = max(max_length + 2, 12)
            ws.column_dimensions[col_letter].width = adjusted_width
    wb.save(output_excel)

    files.download(output_excel)
    print("\n✅ Análisis completado. Archivo: Informe_Modelos_de_Decision_2026.xlsx")
